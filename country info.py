import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Search History File Setup

HISTORY_FILE = "search_history.xlsx"


def create_history_file():
    if not os.path.exists(HISTORY_FILE):

        wb = Workbook()
        ws = wb.active
        ws.title = "History"

        ws.append([
            "Timestamp",
            "Country",
            "Population",
            "Region"
        ])

        wb.save(HISTORY_FILE)

# Fetch Country Data

def fetch_country(country_name):

    try:

        url = f"https://restcountries.com/v3.1/name/{country_name}"

        response = requests.get(url, timeout=10)

        if response.status_code != 200:
            print("Country not found.")
            return None

        data = response.json()

        if not data:
            print("Empty response received.")
            return None

        country = data[0]

        name = country.get("name", {}).get("common", "N/A")

        capital = country.get("capital", ["N/A"])[0]

        population = country.get("population", 0)

        region = country.get("region", "N/A")

        currency = "N/A"

        if "currencies" in country:
            currency = list(country["currencies"].keys())[0]

        timezone = country.get("timezones", ["N/A"])[0]

        return {
            "name": name,
            "capital": capital,
            "population": population,
            "region": region,
            "currency": currency,
            "timezone": timezone
        }

    except requests.exceptions.ConnectionError:
        print("Internet connection error.")

    except requests.exceptions.Timeout:
        print("Request timed out.")

    except Exception as e:
        print("Error:", e)

    return None

# Save Search History

def save_history(country):

    wb = load_workbook(HISTORY_FILE)
    ws = wb["History"]

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        country["name"],
        country["population"],
        country["region"]
    ])

    wb.save(HISTORY_FILE)

# Search Country

def search_country():

    country_name = input("Enter Country Name: ")

    country = fetch_country(country_name)

    if country:

        print("\nCountry Information")
        print("-------------------")
        print("Country :", country["name"])
        print("Capital :", country["capital"])
        print("Population :", country["population"])
        print("Region :", country["region"])
        print("Currency :", country["currency"])
        print("Timezone :", country["timezone"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Country Info"

        ws.append([
            "Country",
            "Capital",
            "Population",
            "Region",
            "Currency",
            "Timezone"
        ])

        ws.append([
            country["name"],
            country["capital"],
            country["population"],
            country["region"],
            country["currency"],
            country["timezone"]
        ])

        wb.save("country_info.xlsx")

        print("\nSaved to country_info.xlsx")

        save_history(country)


# Multi Country Report

def multi_country_report():

    countries = input(
        "Enter Countries separated by commas: "
    ).split(",")

    country_list = []

    for c in countries:

        country = fetch_country(c.strip())

        if country:
            country_list.append(country)

    if len(country_list) == 0:
        print("No valid countries found.")
        return

    wb = Workbook()

    ws = wb.active
    ws.title = "Countries"

    ws.append([
        "Country",
        "Capital",
        "Population",
        "Region",
        "Currency"
    ])

    for country in country_list:

        ws.append([
            country["name"],
            country["capital"],
            country["population"],
            country["region"],
            country["currency"]
        ])

    # Highest Population using loop

    highest = country_list[0]

    for country in country_list:

        if country["population"] > highest["population"]:
            highest = country

    # Lowest Population using loop

    lowest = country_list[0]

    for country in country_list:

        if country["population"] < lowest["population"]:
            lowest = country

    # Regions Covered

    regions = []

    for country in country_list:

        if country["region"] not in regions:
            regions.append(country["region"])

    summary = wb.create_sheet("Summary")

    summary.append(["Total Countries", len(country_list)])

    summary.append([
        "Highest Population Country",
        highest["name"]
    ])

    summary.append([
        "Lowest Population Country",
        lowest["name"]
    ])

    summary.append([
        "Total Regions Covered",
        len(regions)
    ])

    wb.save("multi_country_report.xlsx")

    print("Report Saved Successfully.")

# View Search History

def view_history():

    if not os.path.exists(HISTORY_FILE):
        print("No history available.")
        return

    wb = load_workbook(HISTORY_FILE)
    ws = wb["History"]

    print("\nSearch History")

    for row in ws.iter_rows(values_only=True):
        print(row)


# Export Summary

def export_summary():

    wb = load_workbook(HISTORY_FILE)
    ws = wb["History"]

    countries = []

    region_count = {}

    for row in ws.iter_rows(min_row=2,
                            values_only=True):

        country = row[1]
        population = row[2]
        region = row[3]

        countries.append({
            "name": country,
            "population": population,
            "region": region
        })

        if region in region_count:
            region_count[region] += 1
        else:
            region_count[region] = 1

    if len(countries) == 0:
        print("No data found.")
        return

    highest = countries[0]

    for country in countries:

        if country["population"] > highest["population"]:
            highest = country

    lowest = countries[0]

    for country in countries:

        if country["population"] < lowest["population"]:
            lowest = country

    print("\nRegion Count")

    for region, count in region_count.items():
        print(region, ":", count)

    print("\nHighest Population Country")
    print(highest["name"])

    print("\nLowest Population Country")
    print(lowest["name"])

    region_input = input(
        "\nEnter Region To Search: "
    )

    print("\nCountries In", region_input)

    for country in countries:

        if country["region"].lower() == region_input.lower():
            print(country["name"])

# Main Menu

create_history_file()

while True:

    print("\n===== MENU =====")
    print("1. Search Country")
    print("2. Multi Country Report")
    print("3. View Search History")
    print("4. Export Summary")
    print("5. Exit")

    choice = input("Enter Choice: ")

    if choice == "1":
        search_country()

    elif choice == "2":
        multi_country_report()

    elif choice == "3":
        view_history()

    elif choice == "4":
        export_summary()

    elif choice == "5":
        print("Thank You")
        break

    else:
        print("Invalid Choice")